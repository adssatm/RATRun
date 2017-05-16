#!/usr/bin/python
#-*-coding: utf-8 -*-
import os
import time
import natsort
import openpyxl
from operator import itemgetter
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment , colors
from version import VERSION
_version_ = VERSION


class ExcelManager():

    def __init__(self):
        self.sheetNames = None
        self.fileName = None

    def open_excel_file(self,filename):
        self.wb = openpyxl.load_workbook(filename)
        self.fileName = filename
        self.sheetNames = self.wb.get_sheet_names()
        return self.wb
    
    def get_sheets_name(self):
        return self.sheetNames

    def get_row_count(self,sheetname):
        sheet = self.wb.get_sheet_by_name(sheetname)
        return sheet.max_row

    def get_column_count(self,sheetname):
        sheet = self.wb.get_sheet_by_name(sheetname)
        return sheet.max_column

    def find_row_by_value(self,sheetname,value):
        sheet = self.wb.get_sheet_by_name(sheetname)
        maxRow = sheet.max_row
        maxColumn = sheet.max_column
        for row_index in range(1, maxRow+1):
            for col_index in range(1, maxColumn+1):
                cell = get_column_letter(col_index)+str(row_index)
                cellvalue = sheet[cell].value
                if str(cellvalue) == str(value):
                    return row_index
                    break

    def find_column_by_value(self,sheetname,value):
        sheet = self.wb.get_sheet_by_name(sheetname)
        maxRow = sheet.max_row
        maxColumn = sheet.max_column
        for row_index in range(1, maxRow+1):
            for col_index in range(1, maxColumn+1):
                cell = get_column_letter(col_index)+str(row_index)
                cellvalue = sheet[cell].value
                if str(cellvalue) == str(value):
                    return col_index
                    break

    def get_cellvalue_by_position(self,sheetname,row,column):
        sheet = self.wb.get_sheet_by_name(sheetname)
        cell = get_column_letter(int(column))+str(row)
        cellvalue = sheet[cell].value
        return str(cellvalue)

    def get_cellvalue_by_row(self,sheetname,row):
        sheet = self.wb.get_sheet_by_name(sheetname)
        maxColumn = sheet.max_column
        data = {}
        idict = 0
        for col_index in range(1,maxColumn):
            idict = idict + 1
            cell = get_column_letter(col_index)+str(row)
            cellvalue = sheet[cell].value
            data[idict] = str(cellvalue)
        return data

    def get_cellvalue_by_column(self,sheetname,column):
        sheet = self.wb.get_sheet_by_name(sheetname)
        maxRow = sheet.max_row
        data = {}
        idict = 0
        for row_index in range(2,(maxRow+1)):
            idict = idict + 1
            cell = get_column_letter(column)+str(row_index)
            cellvalue = sheet[cell].value
            data[idict] = str(cellvalue)
        return data
  
    def read_excel_file(self,sheetname,includeEmptyCells=True):
        sheet = self.wb.get_sheet_by_name(sheetname)
        maxRow = sheet.max_row
        maxColumn = sheet.max_column
        data = {}
        for row_index in range(1,maxRow):
            for col_index in range(1,maxColumn):
                cell = get_column_letter(col_index)+str(row_index)
                value = sheet[cell].value
                data[cell] = str(value)
        if includeEmptyCells is True:
            sortedData = natsort.natsorted(data.items(), key=itemgetter(0))
            return sortedData
        else:
            data = dict([(k, v) for (k, v) in data.items() if v])
            OrderedData = natsort.natsorted(data.items(), key=itemgetter(0))
            return OrderedData

    def get_nextcellvalue_by_condition(self,sheetname,value):
        sheet = self.wb.get_sheet_by_name(sheetname)
        maxRow = sheet.max_row
        maxColumn = sheet.max_column
        for row in range(1, maxRow + 1):
            for column in range(1, maxColumn):
                scol = get_column_letter(column)
                cellvalue  = sheet[scol + str(row)].value
                if value == cellvalue:
                    scol = get_column_letter(column+1)
                    return str(sheet[scol + str(row)].value)

    def create_excel_file(self, filenamewithpath, sheetname):
        if filenamewithpath != "" and sheetname != "":
            try:
                book = openpyxl.Workbook()
                arrorigisheet = book.get_sheet_names()
                for isheet in range(0,len(arrorigisheet)):
                    delsheet = book.get_sheet_by_name(arrorigisheet[isheet])
                    book.remove_sheet(delsheet)

                arrnewsheet = str.split(sheetname, ",")
                for isheet in range(0, len(arrnewsheet)):
                    book.create_sheet(arrnewsheet[isheet], isheet)

                book.save(filenamewithpath)
                return True
            except IOError:
                return False

    def write_excel_value(self, sheetname, row, col, val, fontcolor, bgcolor):
        sheet = self.wb.get_sheet_by_name(sheetname)
        try:
            if str.upper(fontcolor) == "RED":
                fontstly = Font(color="E60000")
            elif str.upper(fontcolor) == "BLUE":
                fontstly = Font(color="0000CC")
            elif str.upper(fontcolor) == "GREEN":
                fontstly = Font(color="006600")
            elif str.upper(fontcolor) == "BLACK":
                fontstly = Font(color="0D0D0D")
            else:
                fontstly = Font(color=str.upper(fontcolor))

            if str.upper(bgcolor) != "" and str.upper(bgcolor) != "NOFILL":
                fillcolor = PatternFill("solid", fgColor=str.upper(bgcolor))
            elif str.upper(bgcolor) == "NOFILL":
                fillcolor = "NOFILL"
            else:
                fillcolor = PatternFill("solid", fgColor=colors.YELLOW)

            cell = sheet.cell(row=int(row), column=int(col))
            cell.font = fontstly
            cell.value = val.decode('cp874')

            if fillcolor != "NOFILL":
                cell.fill = fillcolor

            self.wb.save(self.fileName)

            return True
        except:
            print "can not write value on cell at row " + str(row) + " column " + str(col)
            return False